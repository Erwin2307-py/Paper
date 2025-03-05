import streamlit as st
import requests
import feedparser
import pandas as pd
import os
import io
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
import base64

# Google Scholar-Scraping:
try:
    from scholarly import scholarly
except ImportError:
    st.error("Bitte installiere 'scholarly', z.B. via: pip install scholarly")

# ChatGPT/OpenAI:
import openai
# Wichtig: Den API-Key NICHT hartkodieren, sondern st.secrets["openai_api_key"] nutzen.

# PDF-Erzeugung mit fpdf
try:
    from fpdf import FPDF
except ImportError:
    st.error("Bitte installiere 'fpdf', z.B. mit: pip install fpdf")


###############################################################################
# A) ChatGPT: Paper erstellen & lokal speichern
###############################################################################
def generate_paper_via_chatgpt(prompt_text, model="gpt-3.5-turbo"):
    """
    Ruft die ChatGPT-API auf und erzeugt ein Paper (Text).
    Holt den API-Key aus st.secrets["openai_api_key"].
    """
    try:
        openai.api_key = st.secrets["openai_api_key"]  # <-- KEY aus Secrets
        response = openai.ChatCompletion.create(
            model=model,
            messages=[{"role": "user", "content": prompt_text}],
            max_tokens=1200,
            temperature=0.7
        )
        content = response.choices[0].message.content
        return content
    except Exception as e:
        st.error(f"Fehler bei ChatGPT-API: {e}")
        return ""


def save_text_as_pdf(text, pdf_path, title="Paper"):
    """
    Speichert den gegebenen Text in ein PDF (lokal).
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Arial", size=12)

    # Überschrift
    pdf.cell(0, 10, title, ln=1)
    pdf.ln(5)

    lines = text.split("\n")
    for line in lines:
        pdf.multi_cell(0, 8, line)
        pdf.ln(2)

    pdf.output(pdf_path, "F")


###############################################################################
# B) arXiv-Suche & Download & lokal speichern
###############################################################################
def search_arxiv_papers(query, max_results=5):
    base_url = "http://export.arxiv.org/api/query?"
    params = {
        "search_query": f"all:{query}",
        "start": 0,
        "max_results": max_results
    }
    try:
        response = requests.get(base_url, params=params, timeout=10)
        response.raise_for_status()
    except Exception as e:
        st.error(f"Fehler beim Abrufen der Daten von arXiv: {e}")
        return []
    
    feed = feedparser.parse(response.text)
    papers_info = []
    for entry in feed.entries:
        title = entry.title
        summary = entry.summary
        link_pdf = None
        for link in entry.links:
            if link.rel == "related" and "pdf" in link.type:
                link_pdf = link.href
                break
            elif link.type == "application/pdf":
                link_pdf = link.href
                break
        
        papers_info.append({
            "title": title,
            "summary": summary,
            "pdf_url": link_pdf
        })
    return papers_info


def sanitize_filename(fname):
    """Ersetzt unerlaubte Zeichen durch Unterstriche."""
    return re.sub(r"[^a-zA-Z0-9_\-]+", "_", fname)


def download_arxiv_pdf(pdf_url, local_filepath):
    try:
        r = requests.get(pdf_url, timeout=15)
        r.raise_for_status()
        with open(local_filepath, "wb") as f:
            f.write(r.content)
        return True
    except Exception as e:
        st.error(f"Fehler beim Herunterladen der PDF: {e}")
        return False


###############################################################################
# C) Multi-API-Suche (PubMed, Europe PMC, Google Scholar, Semantic Scholar, OpenAlex)
###############################################################################
def flatten_dict(d, parent_key="", sep="__"):
    """Wandelt ein verschachteltes Dict in ein flaches um."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


# --- PubMed-Funktionen ---
def esearch_pubmed(query: str, max_results=100, timeout=10):
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {
        "db": "pubmed",
        "term": query,
        "retmode": "json",
        "retmax": max_results
    }
    try:
        r = requests.get(url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return data.get("esearchresult", {}).get("idlist", [])
    except Exception as e:
        st.error(f"PubMed-Suche fehlgeschlagen: {e}")
        return []


def parse_efetch_response(xml_text: str) -> dict:
    root = ET.fromstring(xml_text)
    pmid_abstract_map = {}
    for article in root.findall(".//PubmedArticle"):
        pmid_el = article.find(".//PMID")
        pmid_val = pmid_el.text if pmid_el is not None else None
        abstract_el = article.find(".//AbstractText")
        abstract_text = abstract_el.text if abstract_el is not None else "n/a"
        if pmid_val:
            pmid_abstract_map[pmid_val] = abstract_text
    return pmid_abstract_map


def fetch_pubmed_abstracts(pmids, timeout=10):
    """Holt Abstracts per efetch."""
    if not pmids:
        return {}
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {
        "db": "pubmed",
        "id": ",".join(pmids),
        "retmode": "xml"
    }
    try:
        r = requests.get(url, params=params, timeout=timeout)
        r.raise_for_status()
        return parse_efetch_response(r.text)
    except Exception as e:
        st.error(f"Fehler beim Abrufen der Abstracts via EFetch: {e}")
        return {}


def get_pubmed_details(pmids: list):
    if not pmids:
        return []
    url_summary = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
    params_sum = {
        "db": "pubmed",
        "id": ",".join(pmids),
        "retmode": "json"
    }
    try:
        r_sum = requests.get(url_summary, params=params_sum, timeout=10)
        r_sum.raise_for_status()
        data_summary = r_sum.json()
    except Exception as e:
        st.error(f"Fehler beim Abrufen von PubMed-Daten (ESummary): {e}")
        return []

    abstracts_map = fetch_pubmed_abstracts(pmids, timeout=10)

    results = []
    for pmid in pmids:
        info = data_summary.get("result", {}).get(pmid, {})
        if not info or pmid == "uids":
            continue

        pubdate = info.get("pubdate", "n/a")
        pubyear = pubdate[:4] if len(pubdate) >= 4 else "n/a"
        doi = info.get("elocationid", "n/a")
        title = info.get("title", "n/a")
        abs_text = abstracts_map.get(pmid, "n/a")
        publisher = info.get("fulljournalname") or info.get("source") or "n/a"

        full_data = dict(info)
        full_data["abstract"] = abs_text

        results.append({
            "Source": "PubMed",
            "Title": title,
            "PubMed ID": pmid,
            "Abstract": abs_text,
            "DOI": doi,
            "Year": pubyear,
            "Publisher": publisher,
            "Population": "n/a",
            "FullData": full_data
        })
    return results


def search_pubmed(query: str, max_results=100):
    pmids = esearch_pubmed(query, max_results=max_results)
    if not pmids:
        return []
    return get_pubmed_details(pmids)


# --- Europe PMC ---
def search_europe_pmc(query: str, max_results=100, timeout=30, retries=3, delay=5):
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {
        "query": query,
        "format": "json",
        "pageSize": max_results
    }
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            data = r.json()
            results = []
            for item in data.get("resultList", {}).get("result", []):
                pub_year = str(item.get("pubYear", "n/a"))
                abstract_text = item.get("abstractText", "n/a")
                publisher = "n/a"
                jinfo = item.get("journalInfo", {})
                if isinstance(jinfo, dict):
                    publisher = jinfo.get("journal", "n/a")

                results.append({
                    "Source": "Europe PMC",
                    "Title": item.get("title", "n/a"),
                    "PubMed ID": item.get("pmid", "n/a"),
                    "Abstract": abstract_text,
                    "DOI": item.get("doi", "n/a"),
                    "Year": pub_year,
                    "Publisher": publisher,
                    "Population": "n/a",
                    "FullData": dict(item)
                })
            return results
        except requests.exceptions.ReadTimeout:
            st.warning(f"Europe PMC: Read Timeout (Versuch {attempt+1}/{retries}). {delay}s warten ...")
            time.sleep(delay)
        except requests.exceptions.RequestException as e:
            st.error(f"Europe PMC-Suche fehlgeschlagen: {e}")
            return []
    st.error("Europe PMC-Suche wiederholt fehlgeschlagen (Timeout).")
    return []


# --- Google Scholar ---
def search_google_scholar(query: str, max_results=100):
    results = []
    try:
        # from scholarly import scholarly  # Schon global importiert
        search_results = scholarly.search_pubs(query)
        for _ in range(max_results):
            publication = next(search_results, None)
            if not publication:
                break
            bib = publication.get('bib', {})
            title = bib.get('title', 'n/a')
            pub_year = bib.get('pub_year', 'n/a')
            abstract = bib.get('abstract', 'n/a')

            results.append({
                "Source": "Google Scholar",
                "Title": title,
                "PubMed ID": "n/a",
                "Abstract": abstract,
                "DOI": "n/a",
                "Year": str(pub_year),
                "Publisher": "n/a",
                "Population": "n/a",
                "FullData": dict(publication)
            })
    except Exception as e:
        st.error(f"Fehler bei der Google Scholar-Suche: {e}")
    return results


# --- Semantic Scholar ---
def search_semantic_scholar(query: str, max_results=100, retries=3, delay=5):
    base_url = "https://api.semanticscholar.org/graph/v1/paper/search"
    params = {
        "query": query,
        "limit": max_results,
        "fields": "title,authors,year,abstract"
    }
    attempt = 0
    results = []
    while attempt < retries:
        try:
            r = requests.get(base_url, params=params, timeout=10)
            r.raise_for_status()
            data = r.json()
            papers = data.get("data", [])
            for paper in papers:
                year_ = str(paper.get("year", "n/a"))
                abstract_ = paper.get("abstract", "n/a")

                results.append({
                    "Source": "Semantic Scholar",
                    "Title": paper.get("title", "n/a"),
                    "PubMed ID": "n/a",
                    "Abstract": abstract_,
                    "DOI": "n/a",
                    "Year": year_,
                    "Publisher": "n/a",
                    "Population": "n/a",
                    "FullData": dict(paper)
                })
            return results
        except requests.exceptions.HTTPError as e:
            if r.status_code == 429:
                st.warning(f"Rate limit bei Semantic Scholar erreicht, warte {delay} Sekunden ...")
                time.sleep(delay)
                attempt += 1
                continue
            else:
                st.error(f"Fehler bei der Semantic Scholar-Suche: {e}")
                return []
        except Exception as e:
            st.error(f"Fehler bei der Semantic Scholar-Suche: {e}")
            return []
    st.error("Semantic Scholar: Rate limit überschritten.")
    return results


# --- OpenAlex ---
def search_openalex(query: str, max_results=100):
    base_url = "https://api.openalex.org/works"
    params = {
        "search": query,
        "per-page": max_results
    }
    try:
        r = requests.get(base_url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        works = data.get("results", [])
        results = []
        for work in works:
            title = work.get("display_name", "n/a")
            publication_year = str(work.get("publication_year", "n/a"))
            doi = work.get("doi", "n/a")
            abstract = "n/a"
            publisher = "n/a"
            full_data = dict(work)

            results.append({
                "Source": "OpenAlex",
                "Title": title,
                "PubMed ID": "n/a",
                "Abstract": abstract,
                "DOI": doi,
                "Year": publication_year,
                "Publisher": publisher,
                "Population": "n/a",
                "FullData": full_data
            })
        return results
    except Exception as e:
        st.error(f"Fehler bei der OpenAlex-Suche: {e}")
        return []


###############################################################################
# D) ChatGPT-Summary von Paper -> PDF
###############################################################################
def create_gpt_paper_pdf(gpt_text, output_stream, title="ChatGPT-Paper"):
    """ Baut ein PDF aus dem ChatGPT-Text. """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Arial", size=12)

    pdf.cell(0, 10, title, ln=1)
    pdf.ln(5)

    lines = gpt_text.split("\n")
    for line in lines:
        pdf.multi_cell(0, 8, line)
        pdf.ln(2)

    pdf_str = pdf.output(dest='S')
    pdf_bytes = pdf_str.encode("latin-1", "replace")
    output_stream.write(pdf_bytes)


###############################################################################
# E) "Profiles" laden (optional)
###############################################################################
def load_settings(profile_name: str):
    if "profiles" in st.session_state:
        profiles = st.session_state["profiles"]
        if profile_name in profiles:
            return profiles[profile_name]
    return None


###############################################################################
# F) Helper, um Excel-kompatible Werte zu erzeugen (wegen ValueError)
###############################################################################
def safe_excel_value(value):
    """
    Konvertiert einen Wert in einen Excel-kompatiblen Typ (str, float, int, bool, None).
    """
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    # Alles andere -> String
    return str(value)


###############################################################################
# G) Modul: Multi-API-Suche + ChatGPT-Download
###############################################################################
def module_codewords_pubmed():
    st.title("Multi-API-Suche (PubMed, Europe PMC, Google Scholar, ...)")

    # Profile (optional)
    if "profiles" not in st.session_state or not st.session_state["profiles"]:
        st.warning("Keine Profile (optional).")
        return

    prof_names = list(st.session_state["profiles"].keys())
    chosen_profile = st.selectbox("Wähle ein Profil:", ["(kein)"] + prof_names)
    if chosen_profile == "(kein)":
        st.info("Kein Profil gewählt.")
        return

    profile_data = load_settings(chosen_profile)
    st.json(profile_data)

    codewords_str = st.text_input("Codewörter:", "")
    logic_option = st.radio("Logik:", ["AND","OR"], 1)

    if st.button("Suche starten"):
        raw_list = [w.strip() for w in codewords_str.replace(",", " ").split() if w.strip()]
        if not raw_list:
            st.warning("Bitte mind. 1 Wort eingeben.")
            return
        query_str = " AND ".join(raw_list) if logic_option == "AND" else " OR ".join(raw_list)
        st.write("Anfrage:", query_str)

        results_all = []

        if profile_data.get("use_pubmed", False):
            pm = search_pubmed(query_str, max_results=100)
            st.write(f"PubMed: {len(pm)}")
            results_all.extend(pm)

        if profile_data.get("use_epmc", False):
            ep = search_europe_pmc(query_str, max_results=100)
            st.write(f"Europe PMC: {len(ep)}")
            results_all.extend(ep)

        if profile_data.get("use_google", False):
            gg = search_google_scholar(query_str, max_results=100)
            st.write(f"Google Scholar: {len(gg)}")
            results_all.extend(gg)

        if profile_data.get("use_semantic", False):
            se = search_semantic_scholar(query_str, max_results=100)
            st.write(f"Semantic Scholar: {len(se)}")
            results_all.extend(se)

        if profile_data.get("use_openalex", False):
            oa = search_openalex(query_str, max_results=100)
            st.write(f"OpenAlex: {len(oa)}")
            results_all.extend(oa)

        if not results_all:
            st.info("Nichts gefunden.")
            return

        st.session_state["search_results"] = results_all

    # Nach der Suche:
    if "search_results" in st.session_state and st.session_state["search_results"]:
        results_all = st.session_state["search_results"]
        df_main = pd.DataFrame([{
            "Title": p.get("Title", "n/a"),
            "PubMed ID": p.get("PubMed ID", "n/a"),
            "Year": p.get("Year", "n/a"),
            "Publisher": p.get("Publisher", "n/a"),
            "Population": p.get("Population", "n/a"),
            "Source": p.get("Source", "n/a"),
            "Abstract": p.get("Abstract", "n/a")
        } for p in results_all])
        st.dataframe(df_main)

        st.subheader("Filter")
        adv_filter = st.checkbox("Filter aktivieren?")

        if adv_filter:
            txt_filter = st.text_input("Suchwort (Title/Publisher):", "")
            years = ["Alle"] + [str(y) for y in range(1900, 2026)]
            y_choice = st.selectbox("Jahr:", years)

            df_filt = df_main.copy()
            if txt_filter.strip():
                tfl = txt_filter.lower()

                def match_f(r):
                    t = (r["Title"] or "").lower()
                    pub = (r["Publisher"] or "").lower()
                    return (tfl in t) or (tfl in pub)

                df_filt = df_filt[df_filt.apply(match_f, axis=1)]
            if y_choice != "Alle":
                df_filt = df_filt[df_filt["Year"] == y_choice]

            st.dataframe(df_filt)

            # Anzeige in neuem Browser-Tab
            if st.button("Gefilterte in neuem Browserfenster anschauen"):
                if df_filt.empty:
                    st.warning("Keine gefilterten Daten.")
                else:
                    html_code = df_filt.to_html(index=False, escape=False)
                    b64_code = base64.b64encode(html_code.encode()).decode()
                    link_html = f'<a href="data:text/html;base64,{b64_code}" target="_blank">Gefilterte Paper öffnen</a>'
                    st.markdown(link_html, unsafe_allow_html=True)

            # Abstract-PDF-Download (ZIP)
            df_filt["identifier"] = df_filt.apply(lambda x: f"{x['Title']}||{x['PubMed ID']}", axis=1)
            chosen_ids = st.multiselect("Paper für Abstract-PDF (ZIP):", df_filt["identifier"].tolist())

            if st.button("Abstract-PDF (ZIP) herunterladen"):
                if not chosen_ids:
                    st.warning("Keine Auswahl.")
                else:
                    selected_papers = []
                    for cid in chosen_ids:
                        row = df_filt.loc[df_filt["identifier"] == cid].iloc[0]
                        match_ = None
                        for rp in results_all:
                            if rp["Title"] == row["Title"] and rp["PubMed ID"] == row["PubMed ID"]:
                                match_ = rp
                                break
                        if match_:
                            selected_papers.append(match_)

                    if not selected_papers:
                        st.warning("Nichts gefunden.")
                    else:
                        zip_buf = io.BytesIO()
                        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                            for sp in selected_papers:
                                fn = sanitize_filename(sp["Title"][:50]) + ".pdf"
                                pdf_io = io.BytesIO()
                                # Erstelle Abstract-PDF
                                pdf = FPDF()
                                pdf.add_page()
                                pdf.set_auto_page_break(auto=True, margin=15)
                                pdf.set_font("Arial", size=12)

                                pdf.cell(0, 10, f"Paper: {sp.get('Title', 'n/a')}", ln=1)
                                pdf.ln(2)
                                pdf.multi_cell(0, 8, sp.get('Abstract', 'n/a'))
                                pdf_str = pdf.output(dest='S')
                                pdf_bytes = pdf_str.encode('latin-1', 'replace')
                                pdf_io.write(pdf_bytes)

                                zf.writestr(fn, pdf_io.getvalue())
                        zip_buf.seek(0)
                        zip_name = f"abstracts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                        st.download_button(
                            "Abstract-PDF-ZIP",
                            data=zip_buf.getvalue(),
                            file_name=zip_name,
                            mime="application/octet-stream"
                        )

            # ChatGPT-Zusammenfassung => PDF
            st.subheader("ChatGPT-Paper Download (ZIP)")
            chosen_ids_gpt = st.multiselect("Paper für ChatGPT-Paper (ZIP):", df_filt["identifier"].tolist())
            if st.button("ChatGPT-Paper (ZIP) erstellen"):
                if not chosen_ids_gpt:
                    st.warning("Keine Auswahl.")
                else:
                    selected_gpt = []
                    for cid in chosen_ids_gpt:
                        row = df_filt.loc[df_filt["identifier"] == cid].iloc[0]
                        found_ = None
                        for rp in results_all:
                            if rp["Title"] == row["Title"] and rp["PubMed ID"] == row["PubMed ID"]:
                                found_ = rp
                                break
                        if found_:
                            selected_gpt.append(found_)

                    if not selected_gpt:
                        st.warning("Nichts gefunden.")
                    else:
                        gpt_zip_buf = io.BytesIO()
                        with zipfile.ZipFile(gpt_zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                            for pap in selected_gpt:
                                prompt_ = (
                                    f"Erstelle ein zusammenfassendes Paper zu folgendem Artikel:\n"
                                    f"Titel: {pap['Title']}\n"
                                    f"Abstract: {pap['Abstract']}\n\n"
                                    f"Gehe auf mögliche Methoden, Ergebnisse, Diskussion ein.\n"
                                )
                                chat_text = generate_paper_via_chatgpt(prompt_)
                                if chat_text:
                                    pdf_mem = io.BytesIO()
                                    create_gpt_paper_pdf(
                                        chat_text,
                                        pdf_mem,
                                        title="ChatGPT-Paper zu: " + pap["Title"][:30]
                                    )
                                    fpdf_name = sanitize_filename(pap["Title"][:50]) + "_gpt.pdf"
                                    zf.writestr(fpdf_name, pdf_mem.getvalue())
                        gpt_zip_buf.seek(0)
                        zipfile_name = f"chatgpt_papers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                        st.download_button(
                            "ChatGPT-Paper-ZIP herunterladen",
                            data=gpt_zip_buf.getvalue(),
                            file_name=zipfile_name,
                            mime="application/octet-stream"
                        )

        # Excel-Export
        st.write("---")
        st.subheader("Excel-Export (alle Paper)")
        codewords_ = "Suchbegriffe"
        stamp_ = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_fn = f"{codewords_}_{stamp_}.xlsx"

        from openpyxl import Workbook

        df_excel_main = df_main[["Title", "PubMed ID", "Abstract", "Year", "Publisher", "Population", "Source"]]
        group_ = defaultdict(list)
        for p in results_all:
            group_[p.get("Source", "n/a")].append(p)

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Main"

        main_cols = ["Title", "PubMed ID", "Abstract", "Year", "Publisher", "Population", "Source"]
        ws_main.append(main_cols)

        def safe_row(row):
            return [safe_excel_value(x) for x in row]

        for idx, row in df_excel_main.iterrows():
            arr = [row[c] for c in main_cols]
            ws_main.append(safe_row(arr))

        # Pro Source ein eigenes Sheet
        for s_name, plist in group_.items():
            wsx = wb.create_sheet(title=s_name[:31])
            all_dicts = []
            for p in plist:
                fd = flatten_dict(p.get("FullData", {}))
                fd["Title"] = p.get("Title", "n/a")
                fd["PubMed ID"] = p.get("PubMed ID", "n/a")
                fd["Abstract"] = p.get("Abstract", "n/a")
                fd["Year"] = p.get("Year", "n/a")
                fd["Publisher"] = p.get("Publisher", "n/a")
                fd["Population"] = p.get("Population", "n/a")
                fd["Source"] = p.get("Source", "n/a")
                all_dicts.append(fd)

            if all_dicts:
                colset = set()
                for d in all_dicts:
                    colset.update(d.keys())
                colnames = sorted(list(colset))

                wsx.append(safe_row(colnames))
                for d in all_dicts:
                    row_ = [safe_excel_value(d.get(c, "")) for c in colnames]
                    wsx.append(row_)

        wb.save(excel_fn)
        with open(excel_fn, "rb") as f:
            st.download_button(
                "Excel herunterladen",
                data=f,
                file_name=excel_fn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


###############################################################################
# H) Haupt-App
###############################################################################
def main():
    st.title("Kombinierte App: ChatGPT-Paper, arXiv-Suche, Multi-API-Suche")

    # Beispiel: Profile definieren (falls nicht existieren)
    if "profiles" not in st.session_state:
        st.session_state["profiles"] = {
            "DefaultProfile": {
                "use_pubmed": True,
                "use_epmc": True,
                "use_google": True,
                "use_semantic": True,
                "use_openalex": True
            }
        }

    menu = ["ChatGPT-Paper", "arXiv-Suche", "Multi-API-Suche"]
    choice = st.sidebar.selectbox("Navigation", menu)

    if choice == "ChatGPT-Paper":
        st.subheader("1) Paper mit ChatGPT generieren & lokal speichern")
        prompt_txt = st.text_area("Prompt:", "Schreibe ein Paper über KI in der Medizin.")
        local_dir = st.text_input("Zielordner lokal:", "chatgpt_papers")
        if st.button("Paper generieren"):
            text = generate_paper_via_chatgpt(prompt_txt)
            if text:
                if not os.path.exists(local_dir):
                    os.makedirs(local_dir, exist_ok=True)
                time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                pdf_name = f"chatgpt_paper_{time_str}.pdf"
                pdf_path = os.path.join(local_dir, pdf_name)
                save_text_as_pdf(text, pdf_path, title="ChatGPT-Paper")
                st.success(f"Paper gespeichert unter: {pdf_path}")

    elif choice == "arXiv-Suche":
        st.subheader("2) arXiv-Suche & PDF-Download (lokal)")
        query = st.text_input("arXiv Suchbegriff:", "quantum computing")
        num = st.number_input("Anzahl", 1, 50, 5)
        local_dir_arxiv = st.text_input("Ordner für Downloads:", "arxiv_papers")

        if st.button("arXiv-Suche starten"):
            results = search_arxiv_papers(query, max_results=num)
            if not results:
                st.info("Keine Treffer.")
            else:
                st.write(f"Treffer: {len(results)}")
                if not os.path.exists(local_dir_arxiv):
                    os.makedirs(local_dir_arxiv, exist_ok=True)
                for i, paper in enumerate(results, start=1):
                    st.write(f"**{i})** {paper['title']}")
                    st.write(paper['summary'][:300], "...")
                    if paper["pdf_url"]:
                        fname = sanitize_filename(paper["title"][:50]) + ".pdf"
                        path_ = os.path.join(local_dir_arxiv, fname)
                        if st.button(f"PDF herunterladen: {fname}", key=f"arxiv_{i}"):
                            ok_ = download_arxiv_pdf(paper["pdf_url"], path_)
                            if ok_:
                                st.success(f"PDF gespeichert: {path_}")
                    else:
                        st.write("_Kein PDF-Link._")
                    st.write("---")

    else:
        st.subheader("3) Multi-API-Suche (PubMed, Europe PMC, Google Scholar, Semantic Scholar, OpenAlex)")
        module_codewords_pubmed()


if __name__ == "__main__":
    st.set_page_config(layout="wide")
    # Stelle sicher, dass openai.api_key in st.secrets["openai_api_key"] existiert.
    main()

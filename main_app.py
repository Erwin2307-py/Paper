import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import re
import datetime
import sys
import concurrent.futures
import os
import PyPDF2
import openai
import time
import json
import pdfplumber
import io
from typing import Dict, Any, Optional
from dotenv import load_dotenv
from PIL import Image
from scholarly import scholarly

# Neu: Excel / openpyxl-Import
import openpyxl

# Neuer Import für die Übersetzung mit google_trans_new (wird aber nur als Fallback benutzt)
from google_trans_new import google_translator

# ------------------------------------------------------------------
# Umgebungsvariablen laden (für OPENAI_API_KEY, falls vorhanden)
# ------------------------------------------------------------------
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# ------------------------------------------------------------------
# Streamlit-Konfiguration
# ------------------------------------------------------------------
st.set_page_config(page_title="Streamlit Multi-Modul Demo", layout="wide")

# ------------------------------------------------------------------
# Login-Funktionalität
# ------------------------------------------------------------------
def login():
    st.title("Login")
    user_input = st.text_input("Username")
    pass_input = st.text_input("Password", type="password")
    if st.button("Login"):
        if (
            user_input == st.secrets["login"]["username"]
            and pass_input == st.secrets["login"]["password"]
        ):
            st.session_state["logged_in"] = True
        else:
            st.error("Login failed. Please check your credentials!")

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login()
    st.stop()

# ------------------------------------------------------------------
# 1) Allgemeine Hilfsfunktionen
# ------------------------------------------------------------------
def clean_html_except_br(text):
    """Entfernt alle HTML-Tags außer <br>."""
    cleaned_text = re.sub(r'</?(?!br\b)[^>]*>', '', text)
    return cleaned_text

def translate_text_openai(text, source_language, target_language, api_key):
    """
    Übersetzt Text über OpenAI-ChatCompletion (sofern key vorhanden).
    Hier im Skript standardmäßig Deutsch->Englisch,
    kann aber flexibel genutzt werden.
    """
    import openai
    openai.api_key = api_key
    prompt_system = (
        f"You are a translation engine from {source_language} to {target_language} for a biotech company called Novogenia "
        f"that focuses on lifestyle and health genetics and health analyses. The outputs you provide will be used directly as "
        f"the translated text blocks. Please translate as accurately as possible in the context of health and lifestyle reporting. "
        f"If there is no appropriate translation, the output should be 'TBD'. Keep the TAGS and do not add additional punctuation."
    )
    prompt_user = f"Translate the following text from {source_language} to {target_language}:\n'{text}'"
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": prompt_system},
                {"role": "user", "content": prompt_user}
            ],
            temperature=0
        )
        translation = response.choices[0].message.content.strip()
        # Entfernt führende/abschließende Anführungszeichen
        if translation and translation[0] in ["'", '"', "‘", "„"]:
            translation = translation[1:]
            if translation and translation[-1] in ["'", '"']:
                translation = translation[:-1]
        translation = clean_html_except_br(translation)
        return translation
    except Exception as e:
        st.warning("Übersetzungsfehler (ChatCompletion): " + str(e))
        return text

# ------------------------------------------------------------------
# NEU: GenotypeFinder-Klasse, um Genotyp-Frequenzen nach Hardy-Weinberg zu berechnen
# ------------------------------------------------------------------
class GenotypeFinder:
    def __init__(self):
        self.ensembl_server = "https://rest.ensembl.org"
    
    def get_variant_info(self, rs_id: str) -> Optional[Dict[str, Any]]:
        """
        Ruft detaillierte Informationen zu einer Variation von Ensembl ab (inkl. Populationsdaten).
        """
        if not rs_id.startswith("rs"):
            rs_id = "rs" + rs_id
        
        ext = f"/variation/human/{rs_id}?pops=1"
        try:
            r = requests.get(self.ensembl_server + ext, headers={"Content-Type": "application/json"}, timeout=10)
            r.raise_for_status()
            return r.json()
        except Exception:
            return None
    
    def calculate_genotype_frequency(self, data: Dict[str, Any], genotype: str) -> Dict[str, float]:
        """
        Berechnet die Genotypfrequenz basierend auf Allelfrequenzen (Hardy-Weinberg).

        Parameters:
          - data: JSON-Daten von der Ensembl API
          - genotype: Genotyp als String (z.B. 'AA', 'AG', 'GG')

        Returns:
          - Dictionary mit Populationen und geschätzten Genotypfrequenzen.
        """
        if not data or 'populations' not in data:
            return {}
        
        genotype = genotype.upper()
        if len(genotype) != 2:
            return {}
        allele1, allele2 = genotype[0], genotype[1]
        
        results = {}
        # Nur 1000GENOMES Populationen durchgehen
        for population in data['populations']:
            pop_name = population.get('population', '')
            if '1000GENOMES' not in pop_name:
                continue
            
            # Allelfrequenzen für diese Population sammeln
            allele_freqs = {}
            for pop_data in data['populations']:
                if pop_data.get('population') == pop_name:
                    a_ = pop_data.get('allele', '')
                    freq_ = pop_data.get('frequency', 0.0)
                    allele_freqs[a_] = freq_
            
            if allele1 not in allele_freqs or allele2 not in allele_freqs:
                continue
            
            # Hardy-Weinberg-Gleichgewicht
            if allele1 == allele2:
                genotype_freq = allele_freqs[allele1] ** 2
            else:
                genotype_freq = 2 * allele_freqs[allele1] * allele_freqs[allele2]
            
            results[pop_name] = genotype_freq
        
        return results

# ------------------------------------------------------------------
# CORE-API, PubMed, EuropePMC, OpenAlex, Google Scholar, 
# Semantic Scholar – Hilfsklassen & Funktionen (vereinfacht)
# ------------------------------------------------------------------
class CoreAPI:
    def __init__(self, api_key):
        self.base_url = "https://api.core.ac.uk/v3/"
        self.headers = {"Authorization": f"Bearer {api_key}"}

    def search_publications(self, query, filters=None, sort=None, limit=100):
        endpoint = "search/works"
        params = {"q": query, "limit": limit}
        if filters:
            filter_expressions = []
            for key, value in filters.items():
                filter_expressions.append(f"{key}:{value}")
            params["filter"] = ",".join(filter_expressions)
        if sort:
            params["sort"] = sort
        r = requests.get(
            self.base_url + endpoint,
            headers=self.headers,
            params=params,
            timeout=15
        )
        r.raise_for_status()
        return r.json()

def check_core_aggregate_connection(api_key="LmAMxdYnK6SDJsPRQCpGgwN7f5yTUBHF", timeout=15):
    """Check, ob CORE aggregator erreichbar ist."""
    try:
        core = CoreAPI(api_key)
        result = core.search_publications("test", limit=1)
        return "results" in result
    except Exception:
        return False

def search_core_aggregate(query, api_key="LmAMxdYnK6SDJsPRQCpGgwN7f5yTUBHF"):
    """Einfache Suche in CORE aggregator."""
    if not api_key:
        return []
    try:
        core = CoreAPI(api_key)
        raw = core.search_publications(query, limit=100)
        out = []
        results = raw.get("results", [])
        for item in results:
            title = item.get("title", "n/a")
            year = str(item.get("yearPublished", "n/a"))
            journal = item.get("publisher", "n/a")
            out.append({
                "PMID": "n/a",
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"CORE search error: {e}")
        return []

# ------------------------------------------------------------------
# 2) PubMed
# ------------------------------------------------------------------
def check_pubmed_connection(timeout=10):
    """Kurzer Verbindungstest zu PubMed."""
    test_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": "test", "retmode": "json"}
    try:
        r = requests.get(test_url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return "esearchresult" in data
    except Exception:
        return False

def search_pubmed_simple(query):
    """Kurze Suche (nur Titel/Journal/Year) in PubMed."""
    esearch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": query, "retmode": "json", "retmax": 100}
    out = []
    try:
        r = requests.get(esearch_url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        idlist = data.get("esearchresult", {}).get("idlist", [])
        if not idlist:
            return out
        esummary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
        sum_params = {"db": "pubmed", "id": ",".join(idlist), "retmode": "json"}
        r2 = requests.get(esummary_url, params=sum_params, timeout=10)
        r2.raise_for_status()
        summary_data = r2.json().get("result", {})
        for pmid in idlist:
            info = summary_data.get(pmid, {})
            title = info.get("title", "n/a")
            pubdate = info.get("pubdate", "")
            year = pubdate[:4] if pubdate else "n/a"
            journal = info.get("fulljournalname", "n/a")
            out.append({
                "PMID": pmid,
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"Error searching PubMed: {e}")
        return []

def fetch_pubmed_abstract(pmid):
    """Holt den Abstract via efetch für eine gegebene PubMed-ID."""
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {"db": "pubmed", "id": pmid, "retmode": "xml"}
    try:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        abs_text = []
        for elem in root.findall(".//AbstractText"):
            if elem.text:
                abs_text.append(elem.text.strip())
        if abs_text:
            return "\n".join(abs_text)
        else:
            return "(No abstract available)"
    except Exception as e:
        return f"(Error: {e})"

def fetch_pubmed_doi_and_link(pmid: str) -> (str, str):
    """
    Versucht, über PubMed den DOI sowie den Link zum Paper herauszufinden.
    Gibt (doi, pubmed_link) zurück. Falls kein DOI gefunden, return ("n/a", link).
    """
    if not pmid or pmid == "n/a":
        return ("n/a", "")
    
    link = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
    
    summary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
    params_sum = {"db": "pubmed", "id": pmid, "retmode": "json"}
    try:
        rs = requests.get(summary_url, params=params_sum, timeout=8)
        rs.raise_for_status()
        data = rs.json()
        result_obj = data.get("result", {}).get(pmid, {})
        eloc = result_obj.get("elocationid", "")
        if eloc and eloc.startswith("doi:"):
            doi_ = eloc.split("doi:", 1)[1].strip()
            if doi_:
                return (doi_, link)
    except Exception:
        pass
    
    efetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params_efetch = {"db": "pubmed", "id": pmid, "retmode": "xml"}
    try:
        r_ef = requests.get(efetch_url, params=params_efetch, timeout=8)
        r_ef.raise_for_status()
        root = ET.fromstring(r_ef.content)
        doi_found = "n/a"
        for aid in root.findall(".//ArticleId"):
            id_type = aid.attrib.get("IdType", "")
            if id_type.lower() == "doi":
                if aid.text:
                    doi_found = aid.text.strip()
                    break
        return (doi_found, link)
    except Exception:
        return ("n/a", link)

# ------------------------------------------------------------------
# 3) Europe PMC
# ------------------------------------------------------------------
def check_europe_pmc_connection(timeout=10):
    """Check, ob Europe PMC erreichbar ist."""
    test_url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {"query": "test", "format": "json", "pageSize": 100}
    try:
        r = requests.get(test_url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return "resultList" in data and "result" in data["resultList"]
    except Exception:
        return False

def search_europe_pmc_simple(query):
    """Kurze Suche in Europe PMC."""
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {
        "query": query,
        "format": "json",
        "pageSize": 100,
        "resultType": "core"
    }
    out = []
    try:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        if "resultList" not in data or "result" not in data["resultList"]:
            return out
        results = data["resultList"]["result"]
        for item in results:
            pmid = item.get("pmid", "n/a")
            title = item.get("title", "n/a")
            year = str(item.get("pubYear", "n/a"))
            journal = item.get("journalTitle", "n/a")
            out.append({
                "PMID": pmid if pmid else "n/a",
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"Europe PMC search error: {e}")
        return []

# ------------------------------------------------------------------
# 4) OpenAlex API
# ------------------------------------------------------------------
BASE_URL = "https://api.openalex.org"

def fetch_openalex_data(entity_type, entity_id=None, params=None):
    url = f"{BASE_URL}/{entity_type}"
    if entity_id:
        url += f"/{entity_id}"
    if params is None:
        params = {}
    params["mailto"] = "your_email@example.com"
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        st.error(f"Fehler: {response.status_code} - {response.text}")
        return None

def search_openalex_simple(query):
    """Kurze Version: Liest die rohen Daten, prüft nur, ob was zurückkommt."""
    search_params = {"search": query}
    return fetch_openalex_data("works", params=search_params)

# ------------------------------------------------------------------
# 5) Google Scholar
# ------------------------------------------------------------------
class GoogleScholarSearch:
    def __init__(self):
        self.all_results = []
    def search_google_scholar(self, base_query):
        try:
            search_results = scholarly.search_pubs(base_query)
            for _ in range(5):
                result = next(search_results)
                title = result['bib'].get('title', "n/a")
                authors = result['bib'].get('author', "n/a")
                year = result['bib'].get('pub_year', "n/a")
                url_article = result.get('url_scholarbib', "n/a")
                abstract_text = result['bib'].get('abstract', "")
                self.all_results.append({
                    "Source": "Google Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": "n/a",
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            st.error(f"Fehler bei der Google Scholar-Suche: {e}")

# ------------------------------------------------------------------
# 6) Semantic Scholar
# ------------------------------------------------------------------
def check_semantic_scholar_connection(timeout=10):
    """Verbindungstest zu Semantic Scholar."""
    try:
        url = "https://api.semanticscholar.org/graph/v1/paper/search"
        params = {"query": "test", "limit": 1, "fields": "title"}
        headers = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, params=params, timeout=timeout)
        response.raise_for_status()
        return response.status_code == 200
    except Exception:
        return False

class SemanticScholarSearch:
    def __init__(self):
        self.all_results = []
    def search_semantic_scholar(self, base_query):
        try:
            url = "https://api.semanticscholar.org/graph/v1/paper/search"
            headers = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
            params = {"query": base_query, "limit": 5, "fields": "title,authors,year,abstract,doi,paperId"}
            response = requests.get(url, headers=headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            for paper in data.get("data", []):
                title = paper.get("title", "n/a")
                authors = ", ".join([author.get("name", "") for author in paper.get("authors", [])])
                year = paper.get("year", "n/a")
                doi = paper.get("doi", "n/a")
                paper_id = paper.get("paperId", "")
                abstract_text = paper.get("abstract", "")
                url_article = f"https://www.semanticscholar.org/paper/{paper_id}" if paper_id else "n/a"
                self.all_results.append({
                    "Source": "Semantic Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": "n/a",
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            st.error(f"Semantic Scholar: {e}")

# ------------------------------------------------------------------
# 7) Dummy-Module-Funktionen
# ------------------------------------------------------------------
def module_paperqa2():
    st.subheader("PaperQA2 Module")
    st.write("Dies ist das PaperQA2 Modul. Hier kannst du weitere Einstellungen und Funktionen für PaperQA2 implementieren.")
    question = st.text_input("Bitte gib deine Frage ein:")
    if st.button("Frage absenden"):
        st.write("Antwort: Dies ist eine Dummy-Antwort auf die Frage:", question)

# ------------------------------------------------------------------
# Placeholder-Seiten
# ------------------------------------------------------------------
def page_home():
    st.title("Welcome to the Main Menu")
    st.write("Choose a module in the sidebar to proceed.")
    st.image("Bild1.jpg", caption="Willkommen!", use_container_width=False, width=600)

def page_codewords_pubmed():
    st.title("Codewords & PubMed Settings")
    st.write("Hier könnte das 'module_codewords_pubmed' stehen ...")
    st.write("Placeholder-Inhalt. Anpassungen bei Bedarf einfügen.")

def page_paper_selection():
    st.title("Paper Selection Settings")
    st.write("Define how you want to pick or exclude certain papers. (Dummy placeholder...)")

def page_analysis():
    st.title("Analysis & Evaluation Settings")
    st.write("Set up your analysis parameters, thresholds, etc. (Dummy placeholder...)")

def page_extended_topics():
    st.title("Extended Topics")
    st.write("Access advanced or extended topics for further research. (Dummy placeholder...)")

def page_paperqa2():
    st.title("PaperQA2")
    module_paperqa2()

def page_excel_online_search():
    st.title("Excel Online Search")
    st.write("Hier könnte Code stehen, um Excel-Online-Suchen durchzuführen (Placeholder).")

def page_online_api_filter():
    st.title("Online-API_Filter (Kombiniert)")
    st.write("Hier könntest du aus mehreren APIs filtern usw. (Placeholder).")

# ------------------------------------------------------------------
# PaperAnalyzer-Klasse
# ------------------------------------------------------------------
class PaperAnalyzer:
    def __init__(self, model="gpt-3.5-turbo"):
        self.model = model
    
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extrahiert reinen Text via PyPDF2."""
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text
    
    def analyze_with_openai(self, text, prompt_template, api_key):
        """Hilfsfunktion, um OpenAI per ChatCompletion aufzurufen."""
        import openai
        openai.api_key = api_key
        if len(text) > 15000:
            text = text[:15000] + "..."
        prompt = prompt_template.format(text=text)
        response = openai.ChatCompletion.create(
            model=self.model,
            messages=[
                {"role": "system", "content": "You are an expert at analyzing scientific papers."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=1500
        )
        return response.choices[0].message.content
    
    def summarize(self, text, api_key):
        """
        Erstellt eine Zusammenfassung (auf Deutsch, um das 'split_summary' etc. zu nutzen).
        Danach kann ins Englische übersetzt werden.
        """
        prompt = (
            "Erstelle eine strukturierte Zusammenfassung des folgenden wissenschaftlichen Papers. "
            "Gliedere sie in mindestens vier klar getrennte Abschnitte (z.B. 1. Hintergrund, 2. Methodik, 3. Ergebnisse, 4. Schlussfolgerungen). "
            "Verwende maximal 500 Wörter:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def extract_key_findings(self, text, api_key):
        """Extrahiere die 5 wichtigsten Erkenntnisse (deutsch)."""
        prompt = (
            "Extrahiere die 5 wichtigsten Erkenntnisse aus diesem wissenschaftlichen Paper. "
            "Liste sie mit Bulletpoints auf:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def identify_methods(self, text, api_key):
        """Ermittelt genutzte Methoden und Techniken (deutsch)."""
        prompt = (
            "Identifiziere und beschreibe die im Paper verwendeten Methoden und Techniken. "
            "Gib zu jeder Methode eine kurze Erklärung:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def evaluate_relevance(self, text, topic, api_key):
        """Bewertet die Relevanz zum Thema (Skala 1-10, deutsch)."""
        prompt = (
            f"Bewerte die Relevanz dieses Papers für das Thema '{topic}' auf einer Skala von 1-10. "
            f"Begründe deine Bewertung:\n\n{{text}}"
        )
        return self.analyze_with_openai(text, prompt, api_key)

# ------------------------------------------------------------------
# Hilfsfunktionen für Summaries, Kohorten, etc.
# ------------------------------------------------------------------
def split_summary(summary_text):
    """Versucht 'Ergebnisse' und 'Schlussfolgerungen' aus dem deutschen Summary zu splitten."""
    pattern = re.compile(
        r'(Ergebnisse(?:\:|\s*\n)|Resultate(?:\:|\s*\n))(?P<results>.*?)(Schlussfolgerungen(?:\:|\s*\n)|Fazit(?:\:|\s*\n))(?P<conclusion>.*)',
        re.IGNORECASE | re.DOTALL
    )
    match = pattern.search(summary_text)
    if match:
        ergebnisse = match.group('results').strip()
        schlussfolgerungen = match.group('conclusion').strip()
        return ergebnisse, schlussfolgerungen
    else:
        return summary_text, ""

def parse_cohort_info(summary_text: str) -> dict:
    """
    Parst grobe Infos zur Kohorte (Anzahl Patienten, Herkunft etc.) 
    aus dem DEUTSCHEN Summary. 
    """
    info = {"study_size": "", "origin": ""}
    pattern_both = re.compile(
        r"(\d+)\s*Patient(?:en)?(?:[^\d]+)(\d+)\s*gesunde\s*Kontroll(?:personen)?",
        re.IGNORECASE
    )
    m_both = pattern_both.search(summary_text)
    if m_both:
        p_count = m_both.group(1)
        c_count = m_both.group(2)
        info["study_size"] = f"{p_count} Patienten / {c_count} Kontrollpersonen"
    else:
        pattern_single_p = re.compile(r"(\d+)\s*Patient(?:en)?", re.IGNORECASE)
        m_single_p = pattern_single_p.search(summary_text)
        if m_single_p:
            info["study_size"] = f"{m_single_p.group(1)} Patienten"
    pattern_origin = re.compile(r"in\s*der\s+(\S+)\s+Bevölkerung", re.IGNORECASE)
    m_orig = pattern_origin.search(summary_text)
    if m_orig:
        info["origin"] = m_orig.group(1).strip()
    return info

# ------------------------------------------------------------------
# ChatGPT-Scoring für eine Liste von Papers (Demo)
# ------------------------------------------------------------------
def chatgpt_online_search_with_genes(papers, codewords, genes, top_k=100):
    openai.api_key = st.secrets.get("OPENAI_API_KEY", "")
    if not openai.api_key:
        st.error("Kein 'OPENAI_API_KEY' in st.secrets hinterlegt.")
        return []
    scored_results = []
    total = len(papers)
    progress = st.progress(0)
    status_text = st.empty()
    genes_str = ", ".join(genes) if genes else ""
    for idx, paper in enumerate(papers, start=1):
        current_title = paper.get("Title", "n/a")
        status_text.text(f"Verarbeite Paper {idx}/{total}: {current_title}")
        progress.progress(idx / total)
        title = paper.get("Title", "n/a")
        abstract = paper.get("Abstract", "n/a")
        prompt = f"""
Codewörter: {codewords}
Gene: {genes_str}

Paper:
Titel: {title}
Abstract: {abstract}

Gib mir eine Zahl von 0 bis 100 (Relevanz), wobei sowohl Codewörter als auch Gene berücksichtigt werden.
"""
        try:
            resp = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
                temperature=0
            )
            raw_text = resp.choices[0].message.content.strip()
            match_ = re.search(r'(\d+)', raw_text)
            if match_:
                score = int(match_.group(1))
            else:
                score = 0
        except Exception as e:
            st.error(f"ChatGPT Fehler beim Scoring: {e}")
            score = 0
        new_item = dict(paper)
        new_item["Relevance"] = score
        scored_results.append(new_item)
    status_text.empty()
    progress.empty()
    scored_results.sort(key=lambda x: x["Relevance"], reverse=True)
    return scored_results[:top_k]

# ------------------------------------------------------------------
# Gemeinsamkeiten & Widersprüche aus mehreren Texten
# ------------------------------------------------------------------
def analyze_papers_for_commonalities_and_contradictions(pdf_texts: Dict[str, str], api_key: str, model: str, method_choice: str = "Standard") -> str:
    import openai
    openai.api_key = api_key

    # 1) Claims je Paper extrahieren
    all_claims = {}
    for fname, txt in pdf_texts.items():
        prompt_claims = f"""
Lies den folgenden Ausschnitt eines wissenschaftlichen Papers (maximal 2000 Tokens).
Extrahiere bitte die wichtigsten 3-5 "Aussagen" (Claims), die das Paper aufstellt.
Nutze als Ausgabe ein kompaktes JSON-Format, z.B:
[
  {{"claim": "Aussage 1"}},
  {{"claim": "Aussage 2"}}
]
Text: {txt[:6000]}
"""
        try:
            resp_claims = openai.ChatCompletion.create(
                model=model,
                messages=[{"role": "user", "content": prompt_claims}],
                temperature=0.3,
                max_tokens=700
            )
            raw = resp_claims.choices[0].message.content.strip()
            try:
                claims_list = json.loads(raw)
            except Exception:
                # Falls Parsing scheitert, Notlösung
                claims_list = [{"claim": raw}]
            if not isinstance(claims_list, list):
                claims_list = [claims_list]
            all_claims[fname] = claims_list
        except Exception as e:
            st.error(f"Fehler beim Claims-Extrahieren in {fname}: {e}")
            all_claims[fname] = []

    merged_claims = []
    for fname, cllist in all_claims.items():
        for cobj in cllist:
            ctext = cobj.get("claim", "(leer)")
            merged_claims.append({
                "paper": fname,
                "claim": ctext
            })
    big_input_str = json.dumps(merged_claims, ensure_ascii=False, indent=2)

    # 2) Gemeinsamkeiten + Widersprüche identifizieren
    if method_choice == "ContraCrow":
        final_prompt = f"""
Nutze die ContraCrow-Methodik, um die folgenden Claims (Aussagen) aus mehreren wissenschaftlichen PDF-Papers zu analysieren. 
Die ContraCrow-Methodik fokussiert sich darauf, systematisch Gemeinsamkeiten und klare Widersprüche zu identifizieren.
Bitte identifiziere:
1) Die zentralen gemeinsamen Aussagen, die in den Papers auftreten.
2) Klare Widersprüche zwischen den Aussagen der verschiedenen Papers.

Antworte ausschließlich in folgendem JSON-Format (ohne zusätzliche Erklärungen):
{{
  "commonalities": [
    "Gemeinsamkeit 1",
    "Gemeinsamkeit 2"
  ],
  "contradictions": [
    {{"paperA": "...", "claimA": "...", "paperB": "...", "claimB": "...", "reason": "..." }},
    ...
  ]
}}

Hier die Claims:
{big_input_str}
"""
    else:
        final_prompt = f"""
Hier sind verschiedene Claims (Aussagen) aus mehreren wissenschaftlichen PDF-Papers im JSON-Format.
Bitte identifiziere:
1) Gemeinsamkeiten zwischen den Papers (Wo überschneiden sich die Aussagen?)
2) Mögliche Widersprüche (Welche Aussagen widersprechen sich klar?)

Antworte NUR in folgendem JSON-Format (ohne weitere Erklärungen):
{{
  "commonalities": [
    "Gemeinsamkeit 1",
    "Gemeinsamkeit 2"
  ],
  "contradictions": [
    {{"paperA": "...", "claimA": "...", "paperB": "...", "claimB": "...", "reason": "..." }},
    ...
  ]
}}

Hier die Claims:
{big_input_str}
"""

    try:
        resp_final = openai.ChatCompletion.create(
            model=model,
            messages=[{"role": "user", "content": final_prompt}],
            temperature=0.0,
            max_tokens=1500
        )
        raw2 = resp_final.choices[0].message.content.strip()
        return raw2
    except Exception as e:
        return f"Fehler bei Gemeinsamkeiten/Widersprüche: {e}"

# ------------------------------------------------------------------
# Eigentliche "Analyze Paper"-Seite (inkl. PDF-Upload etc.)
# ------------------------------------------------------------------
def page_analyze_paper():
    st.title("Analyze Paper - Integriert")
    
    if "api_key" not in st.session_state:
        st.session_state["api_key"] = OPENAI_API_KEY or ""
    
    st.sidebar.header("Einstellungen - PaperAnalyzer")
    new_key_value = st.sidebar.text_input("OpenAI API Key", type="password", value=st.session_state["api_key"])
    st.session_state["api_key"] = new_key_value
    
    model = st.sidebar.selectbox(
        "OpenAI-Modell",
        ["gpt-3.5-turbo", "gpt-3.5-turbo-16k", "gpt-4"],
        index=0
    )
    
    analysis_method = st.sidebar.selectbox("Analyse-Methode (Gemeinsamkeiten & Widersprüche)", ["Standard GPT", "ContraCrow"])
    
    compare_mode = st.sidebar.checkbox("Alle Paper gemeinsam vergleichen (Outlier ausschließen)?")
    
    theme_mode = st.sidebar.radio("Hauptthema bestimmen", ["Manuell", "GPT"])
    
    action = st.sidebar.radio(
        "Analyseart",
        ["Zusammenfassung", "Wichtigste Erkenntnisse", "Methoden & Techniken", "Relevanz-Bewertung", "Tabellen & Grafiken"],
        index=0
    )
    
    user_defined_theme = ""
    if theme_mode == "Manuell":
        user_defined_theme = st.sidebar.text_input("Manuelles Hauptthema (bei Compare-Mode)")
    
    topic = st.sidebar.text_input("Thema für Relevanz-Bewertung (falls relevant)")
    
    # Obwohl wir hier eine Auswahl für Sprache hatten, 
    # wollen wir jetzt das Excel immer in Englisch füllen.
    # -> UI kann aber so bestehen bleiben (wenn gewünscht).
    # In der Praxis ignorieren wir 'output_lang' beim Excel-Schreiben.
    output_lang = st.sidebar.selectbox("Ausgabesprache (nur UI)", ["Deutsch", "Englisch", "Portugiesisch", "Serbisch"], index=0)
    
    uploaded_files = st.file_uploader("PDF-Dateien hochladen", type="pdf", accept_multiple_files=True)
    analyzer = PaperAnalyzer(model=model)
    api_key = st.session_state["api_key"]
    
    if "paper_texts" not in st.session_state:
        st.session_state["paper_texts"] = {}
    
    if "relevant_papers_compare" not in st.session_state:
        st.session_state["relevant_papers_compare"] = None
    if "theme_compare" not in st.session_state:
        st.session_state["theme_compare"] = ""

    def do_outlier_logic(paper_map: dict) -> (list, str):
        """Fragt GPT, welche Paper thematisch (zum gemeinsamen Topic) relevant sind und welches Hauptthema erkannt wird."""
        if theme_mode == "Manuell":
            main_theme = user_defined_theme.strip()
            if not main_theme:
                st.error("Bitte ein manuelles Hauptthema eingeben!")
                return ([], "")
            snippet_list = []
            for name, txt_data in paper_map.items():
                snippet = txt_data[:700].replace("\n", " ")
                snippet_list.append(f'{{"filename": "{name}", "snippet": "{snippet}"}}')
            big_snippet = ",\n".join(snippet_list)
            big_input = f"""
Der Nutzer hat folgendes Hauptthema definiert: '{main_theme}'.

Hier sind mehrere Paper in JSON-Form. Entscheide pro Paper, ob es zu diesem Thema passt oder nicht.
Gib mir am Ende ein JSON-Format zurück:

{{
  "theme": "wiederhole das user-defined theme",
  "papers": [
    {{"filename": "...", "relevant": true/false, "reason": "Kurzer Grund"}}
  ]
}}

Nur das JSON, ohne weitere Erklärungen.

[{big_snippet}]
"""
            try:
                openai.api_key = api_key
                scope_resp = openai.ChatCompletion.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "Du checkst Paper-Snippets auf Relevanz zum user-Thema."},
                        {"role": "user", "content": big_input}
                    ],
                    temperature=0.0,
                    max_tokens=1800
                )
                scope_decision = scope_resp.choices[0].message.content
            except Exception as e1:
                st.error(f"GPT-Fehler bei Compare-Mode (Manuell): {e1}")
                return ([], "")
            st.markdown("#### GPT-Ausgabe (Outlier-Check / Manuell):")
            st.code(scope_decision, language="json")
            json_str = scope_decision.strip()
            # Entferne mögliche ```-Umschließungen
            if json_str.startswith("```"):
                json_str = re.sub(r"```[\w]*\n?", "", json_str)
                json_str = re.sub(r"\n?```", "", json_str)
            try:
                data_parsed = json.loads(json_str)
                papers_info = data_parsed.get("papers", [])
            except Exception as parse_e:
                st.error(f"Fehler beim JSON-Parsing: {parse_e}")
                return ([], "")
            st.write(f"**Hauptthema (Manuell)**: {main_theme}")
            relevant_papers_local = []
            st.write("**Paper-Einstufung**:")
            for p in papers_info:
                fname = p.get("filename", "?")
                rel = p.get("relevant", False)
                reason = p.get("reason", "(none)")
                if rel:
                    relevant_papers_local.append(fname)
                    st.success(f"{fname} => relevant. Begründung: {reason}")
                else:
                    st.warning(f"{fname} => NICHT relevant. Begründung: {reason}")
            return (relevant_papers_local, main_theme)
        else:
            snippet_list = []
            for name, txt_data in paper_map.items():
                snippet = txt_data[:700].replace("\n", " ")
                snippet_list.append(f'{{"filename": "{name}", "snippet": "{snippet}"}}')
            big_snippet = ",\n".join(snippet_list)
            big_input = f"""
Hier sind mehrere Paper in JSON-Form. Bitte ermittele das gemeinsame Hauptthema.
Dann antworte mir in folgendem JSON-Format: 
{{
  "main_theme": "Kurzbeschreibung des gemeinsamen Themas",
  "papers": [
    {{"filename":"...","relevant":true/false,"reason":"Kurzer Grund"}}
  ]
}}

Bitte NUR dieses JSON liefern, ohne weitere Erklärungen:

[{big_snippet}]
"""
            try:
                openai.api_key = api_key
                scope_resp = openai.ChatCompletion.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "Du bist ein Assistent, der Paper thematisch filtert."},
                        {"role": "user", "content": big_input}
                    ],
                    temperature=0.0,
                    max_tokens=1800
                )
                scope_decision = scope_resp.choices[0].message.content
            except Exception as e1:
                st.error(f"GPT-Fehler bei Compare-Mode: {e1}")
                return ([], "")
            st.markdown("#### GPT-Ausgabe (Outlier-Check / GPT):")
            st.code(scope_decision, language="json")
            json_str = scope_decision.strip()
            if json_str.startswith("```"):
                json_str = re.sub(r"```[\w]*\n?", "", json_str)
                json_str = re.sub(r"\n?```", "", json_str)
            try:
                data_parsed = json.loads(json_str)
                main_theme = data_parsed.get("main_theme", "No theme extracted.")
                papers_info = data_parsed.get("papers", [])
            except Exception as parse_e:
                st.error(f"Fehler beim JSON-Parsing: {parse_e}")
                return ([], "")
            st.write(f"**Hauptthema (GPT)**: {main_theme}")
            relevant_papers_local = []
            st.write("**Paper-Einstufung**:")
            for p in papers_info:
                fname = p.get("filename", "?")
                rel = p.get("relevant", False)
                reason = p.get("reason", "(none)")
                if rel:
                    relevant_papers_local.append(fname)
                    st.success(f"{fname} => relevant. Begründung: {reason}")
                else:
                    st.warning(f"{fname} => NICHT relevant. Begründung: {reason}")
            return (relevant_papers_local, main_theme)

    # Hochgeladene PDF(s)
    if uploaded_files and api_key:
        # ---------------------- Compare Mode ----------------------
        if compare_mode:
            st.write("### Vergleichsmodus: Outlier-Paper ausschließen")
            if st.button("Vergleichs-Analyse starten"):
                paper_map = {}
                for fpdf in uploaded_files:
                    txt = analyzer.extract_text_from_pdf(fpdf)
                    if txt.strip():
                        paper_map[fpdf.name] = txt
                    else:
                        st.warning(f"Kein Text aus {fpdf.name} extrahierbar (übersprungen).")
                if not paper_map:
                    st.error("Keine verwertbaren Paper.")
                    return
                relevant_papers, discovered_theme = do_outlier_logic(paper_map)
                st.session_state["relevant_papers_compare"] = relevant_papers
                st.session_state["theme_compare"] = discovered_theme
                if not relevant_papers:
                    st.error("Keine relevanten Paper nach Outlier-Check übrig.")
                    return
                combined_text = ""
                for rp in relevant_papers:
                    combined_text += f"\n=== {rp} ===\n{paper_map[rp]}"
                if action == "Tabellen & Grafiken":
                    final_result = "Tabellen & Grafiken nicht im kombinierten Compare-Mode implementiert."
                else:
                    if action == "Zusammenfassung":
                        final_result = analyzer.summarize(combined_text, api_key)
                    elif action == "Wichtigste Erkenntnisse":
                        final_result = analyzer.extract_key_findings(combined_text, api_key)
                    elif action == "Methoden & Techniken":
                        final_result = analyzer.identify_methods(combined_text, api_key)
                    elif action == "Relevanz-Bewertung":
                        if not topic:
                            st.error("Bitte Thema angeben!")
                            return
                        final_result = analyzer.evaluate_relevance(combined_text, topic, api_key)
                    else:
                        final_result = "(Keine Analyseart gewählt.)"
                # Hier nur UI-Ausgabe (optional übersetzen)
                if output_lang != "Deutsch":
                    final_result = translate_text_openai(final_result, "German", output_lang, api_key)
                st.subheader("Ergebnis des Compare-Mode:")
                st.write(final_result)

        # --------------------- Einzel/Multi - ohne Outlier-Check ---------------------
        else:
            st.write("### Einzel- oder Multi-Modus (kein Outlier-Check)")
            pdf_options = ["(Alle)"] + [f"{i+1}) {f.name}" for i, f in enumerate(uploaded_files)]
            selected_pdf = st.selectbox("Wähle eine PDF für Einzel-Analyse oder '(Alle)'", pdf_options)
            
            col_analysis, col_contradiction = st.columns(2)

            with col_analysis:
                if st.button("Analyse starten (Einzel-Modus)"):
                    if selected_pdf == "(Alle)":
                        files_to_process = uploaded_files
                    else:
                        idx = pdf_options.index(selected_pdf) - 1
                        if idx < 0:
                            st.warning("Keine Datei ausgewählt.")
                            return
                        files_to_process = [uploaded_files[idx]]
                    
                    final_result_text = []
                    for fpdf in files_to_process:
                        text_data = ""
                        if action != "Tabellen & Grafiken":
                            with st.spinner(f"Extrahiere Text aus {fpdf.name}..."):
                                text_data = analyzer.extract_text_from_pdf(fpdf)
                                if not text_data.strip():
                                    st.error(f"Kein Text aus {fpdf.name} extrahierbar.")
                                    continue
                                st.success(f"Text aus {fpdf.name} extrahiert!")
                                st.session_state["paper_text"] = text_data[:15000]
                        
                        result = ""
                        if action == "Zusammenfassung":
                            with st.spinner(f"Erstelle Zusammenfassung für {fpdf.name}..."):
                                result = analyzer.summarize(text_data, api_key)
                        elif action == "Wichtigste Erkenntnisse":
                            with st.spinner(f"Extrahiere Erkenntnisse aus {fpdf.name}..."):
                                result = analyzer.extract_key_findings(text_data, api_key)
                        elif action == "Methoden & Techniken":
                            with st.spinner(f"Identifiziere Methoden aus {fpdf.name}..."):
                                result = analyzer.identify_methods(text_data, api_key)
                        elif action == "Relevanz-Bewertung":
                            if not topic:
                                st.error("Bitte Thema angeben!")
                                return
                            with st.spinner(f"Bewerte Relevanz von {fpdf.name}..."):
                                result = analyzer.evaluate_relevance(text_data, topic, api_key)
                        elif action == "Tabellen & Grafiken":
                            with st.spinner(f"Suche Tabellen/Grafiken in {fpdf.name}..."):
                                all_tables_text = []
                                try:
                                    with pdfplumber.open(fpdf) as pdf_:
                                        for page_number, page in enumerate(pdf_.pages, start=1):
                                            st.markdown(f"### Seite {page_number} in {fpdf.name}")
                                            tables = page.extract_tables()
                                            if tables:
                                                st.markdown("**Tabellen auf dieser Seite**")
                                                for table_idx, table_data in enumerate(tables, start=1):
                                                    if not table_data:
                                                        st.write("Leere Tabelle erkannt.")
                                                        continue
                                                    first_row = table_data[0]
                                                    data_rows = table_data[1:]
                                                    if not data_rows:
                                                        st.write("Nur Header vorhanden.")
                                                        data_rows = table_data
                                                        first_row = [f"Col_{i}" for i in range(len(data_rows[0]))]
                                                    import pandas as pd
                                                    new_header = []
                                                    used_cols = {}
                                                    for col in first_row:
                                                        col_str = col if col else "N/A"
                                                        if col_str not in used_cols:
                                                            used_cols[col_str] = 1
                                                            new_header.append(col_str)
                                                        else:
                                                            used_cols[col_str] += 1
                                                            new_header.append(f"{col_str}.{used_cols[col_str]}")
                                                    if any(len(row) != len(new_header) for row in data_rows):
                                                        st.write("Warnung: Inkonsistente Spaltenanzahl in der Tabelle.")
                                                        df = pd.DataFrame(table_data)
                                                    else:
                                                        df = pd.DataFrame(data_rows, columns=new_header)
                                                    st.write(f"**Tabelle {table_idx}** in {fpdf.name}:")
                                                    st.dataframe(df)
                                                    table_str = df.to_csv(index=False)
                                                    all_tables_text.append(f"Seite {page_number} - Tabelle {table_idx}\n{table_str}\n")
                                            else:
                                                st.write("Keine Tabellen auf dieser Seite.")
                                            
                                            images = page.images
                                            if images:
                                                st.markdown("**Bilder/Grafiken auf dieser Seite**")
                                                for img_index, img_dict in enumerate(images, start=1):
                                                    xref = img_dict.get("xref")
                                                    if xref is not None:
                                                        extracted_img = page.extract_image(xref)
                                                        if extracted_img:
                                                            image_data = extracted_img["image"]
                                                            image = Image.open(io.BytesIO(image_data))
                                                            st.write(f"**Bild {img_index}** in {fpdf.name}:")
                                                            st.image(image, use_column_width=True)
                                            else:
                                                st.write("Keine Bilder auf dieser Seite.")
                                    st.markdown(f"### Volltext-Suche 'Table' in {fpdf.name}")
                                    try:
                                        text_all_pages = ""
                                        with pdfplumber.open(fpdf) as pdf2:
                                            for pg in pdf2.pages:
                                                t_ = pg.extract_text() or ""
                                                text_all_pages += t_ + "\n"
                                        lines = text_all_pages.splitlines()
                                        matches = [ln for ln in lines if "Table" in ln]
                                        if matches:
                                            st.write("Zeilen mit 'Table':")
                                            for ln in matches:
                                                st.write(f"- {ln}")
                                        else:
                                            st.write("Keine Erwähnung von 'Table'.")
                                    except Exception as e2:
                                        st.warning(f"Fehler bei Volltext-Suche 'Table': {e2}")
                                    if len(all_tables_text) > 0:
                                        combined_tables_text = "\n".join(all_tables_text)
                                        if len(combined_tables_text) > 14000:
                                            combined_tables_text = combined_tables_text[:14000] + "..."
                                        gpt_prompt = (
                                            "Bitte analysiere die folgenden Tabellen aus einem wissenschaftlichen PDF. "
                                            "Fasse die wichtigsten Erkenntnisse zusammen und gib (wenn möglich) eine "
                                            "kurze Interpretation in Bezug auf Lifestyle und Health Genetics:\n\n"
                                            f"{combined_tables_text}"
                                        )
                                        try:
                                            openai.api_key = api_key
                                            gpt_resp = openai.ChatCompletion.create(
                                                model=model,
                                                messages=[
                                                    {"role": "system", "content": "Du bist ein Experte für PDF-Tabellenanalyse."},
                                                    {"role": "user", "content": gpt_prompt}
                                                ],
                                                temperature=0.3,
                                                max_tokens=1000
                                            )
                                            result = gpt_resp.choices[0].message.content
                                        except Exception as e2:
                                            st.error(f"Fehler bei GPT-Tabellenanalyse: {str(e2)}")
                                            result = "(Fehler bei GPT-Auswertung)"
                                    else:
                                        result = f"In {fpdf.name} keine Tabellen erkannt."
                                except Exception as e_:
                                    st.error(f"Fehler in {fpdf.name}: {str(e_)}")
                                    result = f"(Fehler in {fpdf.name})"
                        
                        # Nur UI-Ausgabe
                        if result and output_lang != "Deutsch" and action != "Tabellen & Grafiken":
                            # Falls gewünscht, die Anzeige anpassen
                            result = translate_text_openai(result, "German", output_lang, api_key)
                        
                        final_result_text.append(f"**Ergebnis für {fpdf.name}:**\n\n{result}")
                    
                    st.subheader("Ergebnis der (Multi-)Analyse (Einzelmodus):")
                    combined_output = "\n\n---\n\n".join(final_result_text)
                    st.markdown(combined_output)

            with col_contradiction:
                st.write("Widerspruchsanalyse (Hochgeladene Paper)")
                if st.button("Widerspruchsanalyse jetzt starten"):
                    if "paper_texts" not in st.session_state or not st.session_state["paper_texts"]:
                        st.session_state["paper_texts"] = {}
                        for upf in uploaded_files:
                            t_ = analyzer.extract_text_from_pdf(upf)
                            if t_.strip():
                                st.session_state["paper_texts"][upf.name] = t_
                    paper_texts = st.session_state["paper_texts"]
                    if not paper_texts:
                        st.error("Keine Texte für die Widerspruchsanalyse vorhanden (hochgeladene Paper).")
                        return
                    with st.spinner("Analysiere hochgeladene Paper auf Gemeinsamkeiten & Widersprüche..."):
                        result_json_str = analyze_papers_for_commonalities_and_contradictions(
                            pdf_texts=paper_texts,
                            api_key=api_key,
                            model=model,
                            method_choice="ContraCrow" if analysis_method == "ContraCrow" else "Standard"
                        )
                        st.subheader("Ergebnis (JSON)")
                        st.code(result_json_str, language="json")
                        try:
                            data_js = json.loads(result_json_str)
                            common = data_js.get("commonalities", [])
                            contras = data_js.get("contradictions", [])
                            st.write("## Gemeinsamkeiten")
                            if common:
                                for c in common:
                                    st.write(f"- {c}")
                            else:
                                st.info("Keine Gemeinsamkeiten erkannt.")
                            st.write("## Widersprüche")
                            if contras:
                                for i, cobj in enumerate(contras, start=1):
                                    st.write(f"Widerspruch {i}:")
                                    st.write(f"- **Paper A**: {cobj.get('paperA')} => {cobj.get('claimA')}")
                                    st.write(f"- **Paper B**: {cobj.get('paperB')} => {cobj.get('claimB')}")
                                    st.write(f"  Grund: {cobj.get('reason','(none)')}")
                            else:
                                st.info("Keine Widersprüche erkannt.")
                        except Exception as e:
                            st.warning(f"Die GPT-Ausgabe konnte nicht als valides JSON geparst werden.\nFehler: {e}")
    else:
        if not api_key:
            st.warning("Bitte OpenAI API-Key eingeben!")
        elif not uploaded_files:
            st.info("Bitte eine oder mehrere PDF-Dateien hochladen!")

    # ---------------- Excel-Ausgabe-Logik (Demo) ----------------
    st.write("---")
    st.write("## Alle Analysen & Excel-Ausgabe (Multi-PDF) (in English)")

    if "excel_downloads" not in st.session_state:
        st.session_state["excel_downloads"] = []

    if uploaded_files and api_key:
        if st.button("Alle Analysen durchführen & in Excel speichern (Multi)"):
            st.session_state["excel_downloads"].clear()
            with st.spinner("Analyzing all uploaded PDFs for Excel..."):
                if compare_mode:
                    if not st.session_state["relevant_papers_compare"]:
                        paper_map_auto = {}
                        for fpdf in uploaded_files:
                            txt = analyzer.extract_text_from_pdf(fpdf)
                            if txt.strip():
                                paper_map_auto[fpdf.name] = txt
                        if not paper_map_auto:
                            st.error("No analyzable PDFs.")
                            return
                        relevant_papers_auto, discovered_theme_auto = do_outlier_logic(paper_map_auto)
                        st.session_state["relevant_papers_compare"] = relevant_papers_auto
                        st.session_state["theme_compare"] = discovered_theme_auto
                    
                    relevant_list_for_excel = st.session_state["relevant_papers_compare"] or []
                    if not relevant_list_for_excel:
                        st.error("No relevant papers after outlier-check.")
                        return
                    selected_files_for_excel = [f for f in uploaded_files if f.name in relevant_list_for_excel]
                else:
                    selected_files_for_excel = uploaded_files

                for fpdf in selected_files_for_excel:
                    text = analyzer.extract_text_from_pdf(fpdf)
                    if not text.strip():
                        st.error(f"No text extracted from {fpdf.name}. Skipping...")
                        continue
                    
                    # 1) Summary (DE)
                    summary_de = analyzer.summarize(text, api_key)
                    # -> Translate to English
                    summary_en = translate_text_openai(summary_de, "German", "English", api_key)
                    
                    # 2) Key findings (DE -> EN)
                    key_findings_de = analyzer.extract_key_findings(text, api_key)
                    key_findings_en = translate_text_openai(key_findings_de, "German", "English", api_key)
                    
                    # 3) Relevance (DE -> EN), if topic given
                    if not topic:
                        relevance_de = "(No topic => no relevance)"
                        relevance_en = relevance_de
                    else:
                        relevance_de = analyzer.evaluate_relevance(text, topic, api_key)
                        relevance_en = translate_text_openai(relevance_de, "German", "English", api_key)
                    
                    # 4) Methods (DE -> EN)
                    methods_de = analyzer.identify_methods(text, api_key)
                    methods_en = translate_text_openai(methods_de, "German", "English", api_key)
                    
                    # 5) Then parse summary for "Ergebnisse" + "Schlussfolgerungen" (in DE),
                    #    afterwards translate to English
                    ergebnisse, schlussfolgerungen = split_summary(summary_de)
                    ergebnisse_en = translate_text_openai(ergebnisse, "German", "English", api_key)
                    schlussfolgerungen_en = translate_text_openai(schlussfolgerungen, "German", "English", api_key)
                    
                    # 6) Cohort info (DE -> EN)
                    cohort_data = parse_cohort_info(summary_de)
                    cohort_info_de = ""
                    if cohort_data["study_size"] or cohort_data["origin"]:
                        c_combo = cohort_data["study_size"]
                        if cohort_data["origin"]:
                            c_combo += ", " + cohort_data["origin"]
                        cohort_info_de = c_combo.strip(", ")
                    cohort_info_en = translate_text_openai(cohort_info_de, "German", "English", api_key)
                    
                    # 7) Gene + rs
                    pattern_obvious = re.compile(r"in the\s+([A-Za-z0-9_-]+)\s+gene", re.IGNORECASE)
                    match_text = re.search(pattern_obvious, text)
                    gene_via_text = match_text.group(1) if match_text else None
                    
                    rs_pat = r"(rs\d+)"
                    found_rs_match = re.search(rs_pat, text)
                    rs_num = found_rs_match.group(1) if found_rs_match else None
                    
                    genotype_regex = r"\b([ACGT]{2,3})\b"
                    lines = text.split("\n")
                    found_pairs = []
                    for line_ in lines:
                        matches_ = re.findall(genotype_regex, line_)
                        if matches_:
                            for m_ in matches_:
                                found_pairs.append((m_, line_.strip()))
                    unique_geno_pairs = []
                    for gp in found_pairs:
                        if gp not in unique_geno_pairs:
                            unique_geno_pairs.append(gp)
                    
                    gf = GenotypeFinder()
                    data_ens = gf.get_variant_info(rs_num) if rs_num else None
                    
                    # 8) Publication year
                    pub_year_match = re.search(r"\b(20[0-9]{2})\b", text)
                    year_for_excel = pub_year_match.group(1) if pub_year_match else "n/a"
                    
                    # 9) PubMed ID
                    pmid_pattern = re.compile(r"\bPMID:\s*(\d+)\b", re.IGNORECASE)
                    pmid_match = pmid_pattern.search(text)
                    pmid_found = pmid_match.group(1) if pmid_match else "n/a"
                    
                    doi_final = "n/a"
                    link_pubmed = ""
                    if pmid_found != "n/a":
                        doi_final, link_pubmed = fetch_pubmed_doi_and_link(pmid_found)

                    # 10) Excel-Vorlage laden
                    try:
                        wb = openpyxl.load_workbook("vorlage_paperqa2.xlsx")
                    except FileNotFoundError:
                        st.error("Template 'vorlage_paperqa2.xlsx' was not found!")
                        return
                    ws = wb.active

                    # A) Main theme (falls relevant)
                    main_theme_for_excel = st.session_state.get("theme_compare", "N/A")
                    if not compare_mode and theme_mode == "Manuell":
                        main_theme_for_excel = user_defined_theme or "N/A"
                    
                    # B) Fill the cells (in English)
                    ws["D2"].value = main_theme_for_excel
                    ws["J2"].value = datetime.datetime.now().strftime("%Y-%m-%d")
                    
                    ws["D5"].value = gene_via_text if gene_via_text else ""
                    ws["D6"].value = rs_num if rs_num else ""
                    
                    # Genotypes (max. 3)
                    max_genos_to_show = 3
                    genotype_entries = unique_geno_pairs[:max_genos_to_show]
                    for i in range(max_genos_to_show):
                        row_i = 10 + i
                        if i < len(genotype_entries):
                            genotype_str = genotype_entries[i][0]  # e.g. "AG"
                            ws[f"D{row_i}"].value = genotype_str
                            
                            if data_ens and genotype_str:
                                genotype_freqs = gf.calculate_genotype_frequency(data_ens, genotype_str)
                                global_freq = genotype_freqs.get("1000GENOMES:phase_3:ALL")
                                if global_freq is not None:
                                    freq_str = f"Global population frequency: {global_freq:.4f}"
                                else:
                                    freq_str = "No global frequency"
                            else:
                                freq_str = "No rsID or no data"
                            ws[f"E{row_i}"].value = freq_str
                        else:
                            ws[f"D{row_i}"] = ""
                            ws[f"E{row_i}"] = ""
                    
                    # year, cohort, key findings, results, conclusion
                    ws["C20"].value = year_for_excel
                    ws["D20"].value = cohort_info_en
                    ws["E20"].value = key_findings_en
                    ws["G21"].value = ergebnisse_en
                    ws["G22"].value = schlussfolgerungen_en
                    
                    # pubmed ID / link / doi
                    ws["J21"].value = pmid_found if pmid_found != "n/a" else ""
                    ws["J22"].value = link_pubmed if link_pubmed else ""
                    ws["I22"].value = doi_final if doi_final != "n/a" else ""
                    
                    # Save to memory
                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    output_buffer.seek(0)
                    
                    xlsx_name = f"analysis_{fpdf.name.replace('.pdf','')}.xlsx"
                    st.session_state["excel_downloads"].append({
                        "label": f"Download Excel for {fpdf.name}",
                        "data": output_buffer.getvalue(),
                        "file_name": xlsx_name
                    })

    # -------------- Download Buttons -------------
    if "excel_downloads" in st.session_state and st.session_state["excel_downloads"]:
        st.write("## Generated Excel Downloads:")
        for dl in st.session_state["excel_downloads"]:
            st.download_button(
                label=dl["label"],
                data=dl["data"],
                file_name=dl["file_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ------------------- ChatGPT-Scoring-Teil (nur Demo) -------------------
    st.write("---")
    st.write("## Einzelanalyse der nach ChatGPT-Scoring ausgewählten Paper")
    
    if st.button("Scoring jetzt durchführen"):
        if "search_results" in st.session_state and st.session_state["search_results"]:
            codewords_str = st.session_state.get("codewords", "")
            selected_genes = st.session_state.get("selected_genes", [])
            scored_list = chatgpt_online_search_with_genes(
                papers=st.session_state["search_results"],
                codewords=codewords_str,
                genes=selected_genes,
                top_k=200
            )
            st.session_state["scored_list"] = scored_list
            st.success("Scored Paper successfully stored in st.session_state['scored_list']!")
        else:
            st.info("No previous search results found, so no scoring possible.")
    
    if "scored_list" not in st.session_state or not st.session_state["scored_list"]:
        st.info("No scored papers yet. Please click 'Scoring jetzt durchführen' first.")
        return
    
    st.subheader("Single Analysis of ChatGPT-Scored Papers")
    scored_titles = [paper["Title"] for paper in st.session_state["scored_list"]]
    chosen_title = st.selectbox(
        "Select a paper from the scoring list:",
        options=["(Bitte wählen)"] + scored_titles
    )
    
    analysis_choice_for_scored_paper = st.selectbox(
        "Which analysis do you want to perform?",
        ["(Keine Auswahl)", "Zusammenfassung", "Wichtigste Erkenntnisse", "Methoden & Techniken", "Relevanz-Bewertung"]
    )
    
    if chosen_title != "(Bitte wählen)":
        selected_paper = next((p for p in st.session_state["scored_list"] if p["Title"] == chosen_title), None)
        if selected_paper:
            st.write("**Title:** ", selected_paper.get("Title", "n/a"))
            st.write("**Source:** ", selected_paper.get("Source", "n/a"))
            st.write("**PubMed ID:** ", selected_paper.get("PubMed ID", "n/a"))
            st.write("**Year:** ", selected_paper.get("Year", "n/a"))
            st.write("**Publisher:** ", selected_paper.get("Publisher", "n/a"))
            st.write("**Abstract:**")
            abstract = selected_paper.get("Abstract") or ""
            if abstract.strip():
                st.markdown(f"> {abstract}")
            else:
                st.warning("No abstract available.")
            
            if st.button("Analyse für dieses Paper durchführen"):
                if not abstract.strip():
                    st.error("No abstract, can't analyze.")
                    return
                if analysis_choice_for_scored_paper == "Zusammenfassung":
                    res_de = analyzer.summarize(abstract, api_key)
                    res_en = translate_text_openai(res_de, "German", "English", api_key)
                    st.write("### Analysis Result (English):")
                    st.write(res_en)
                elif analysis_choice_for_scored_paper == "Wichtigste Erkenntnisse":
                    res_de = analyzer.extract_key_findings(abstract, api_key)
                    res_en = translate_text_openai(res_de, "German", "English", api_key)
                    st.write("### Analysis Result (English):")
                    st.write(res_en)
                elif analysis_choice_for_scored_paper == "Methoden & Techniken":
                    res_de = analyzer.identify_methods(abstract, api_key)
                    res_en = translate_text_openai(res_de, "German", "English", api_key)
                    st.write("### Analysis Result (English):")
                    st.write(res_en)
                elif analysis_choice_for_scored_paper == "Relevanz-Bewertung":
                    if not topic:
                        st.error("Please enter a topic (Sidebar).")
                        return
                    res_de = analyzer.evaluate_relevance(abstract, topic, api_key)
                    res_en = translate_text_openai(res_de, "German", "English", api_key)
                    st.write("### Analysis Result (English):")
                    st.write(res_en)
                else:
                    st.info("No valid analysis selected.")
        else:
            st.warning("Paper not found (unexpected error).")

    st.write("---")
    st.header("PaperQA Multi-Paper Analyzer: Commonalities & Contradictions (Scored Papers)")
    if st.button("Analyse (Gescorte Paper) durchführen"):
        if "scored_list" in st.session_state and st.session_state["scored_list"]:
            paper_texts = {}
            for paper in st.session_state["scored_list"]:
                title = paper.get("Title", "Untitled")
                abstract = paper.get("Abstract") or ""
                if abstract.strip():
                    paper_texts[title] = abstract
                else:
                    st.warning(f"No abstract for {title}.")
            if not paper_texts:
                st.error("No texts for the analysis.")
            else:
                with st.spinner("Analyzing scored papers for commonalities & contradictions..."):
                    result_json_str = analyze_papers_for_commonalities_and_contradictions(
                        paper_texts,
                        api_key,
                        model,
                        method_choice="ContraCrow" if analysis_method == "ContraCrow" else "Standard"
                    )
                    st.subheader("Result (JSON)")
                    st.code(result_json_str, language="json")
                    try:
                        data_js = json.loads(result_json_str)
                        common = data_js.get("commonalities", [])
                        contras = data_js.get("contradictions", [])
                        st.write("## Commonalities")
                        if common:
                            for c in common:
                                st.write(f"- {c}")
                        else:
                            st.info("No commonalities found.")
                        st.write("## Contradictions")
                        if contras:
                            for i, cobj in enumerate(contras, start=1):
                                st.write(f"Contradiction {i}:")
                                st.write(f"- **Paper A**: {cobj.get('paperA')} => {cobj.get('claimA')}")
                                st.write(f"- **Paper B**: {cobj.get('paperB')} => {cobj.get('claimB')}")
                                st.write(f"  Reason: {cobj.get('reason','(none)')}")
                        else:
                            st.info("No contradictions found.")
                    except Exception as e:
                        st.warning("The GPT output could not be parsed as valid JSON.")
        else:
            st.error("No scored papers. Please run scoring first.")

# ------------------------------------------------------------------
# Sidebar & Chat
# ------------------------------------------------------------------
def sidebar_module_navigation():
    st.sidebar.title("Module Navigation")
    pages = {
        "Home": page_home,
        "Online-API_Filter": page_online_api_filter,
        "3) Codewords & PubMed": page_codewords_pubmed,
        "Analyze Paper": page_analyze_paper,
    }
    for label, page in pages.items():
        if st.sidebar.button(label, key=label):
            st.session_state["current_page"] = label
    if "current_page" not in st.session_state:
        st.session_state["current_page"] = "Home"
    return pages.get(st.session_state["current_page"], page_home)

def answer_chat(question: str) -> str:
    """Simple example: Uses any paper-text (if stored in session) + GPT for context."""
    api_key = st.session_state.get("api_key", "")
    paper_text = st.session_state.get("paper_text", "")
    if not api_key:
        return f"(No API Key) Echo: {question}"
    if not paper_text.strip():
        sys_msg = "You are a helpful assistant for general questions."
    else:
        sys_msg = (
            "You are a helpful assistant, and here is a paper as context:\n\n"
            + paper_text[:12000] + "\n\n"
            "Please use it to answer questions as expertly as possible."
        )
    openai.api_key = api_key
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": question}
            ],
            temperature=0.3,
            max_tokens=400
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"OpenAI error: {e}"

def main():
    col_left, col_right = st.columns([4, 1])
    
    with col_left:
        page_fn = sidebar_module_navigation()
        page_fn()
    
    with col_right:
        st.subheader("Chatbot")
        if "chat_history" not in st.session_state:
            st.session_state["chat_history"] = []
        user_input = st.text_input("Your question here", key="chatbot_right_input")
        if st.button("Send (Chat)", key="chatbot_right_send"):
            if user_input.strip():
                st.session_state["chat_history"].append(("user", user_input))
                bot_answer = answer_chat(user_input)
                st.session_state["chat_history"].append(("bot", bot_answer))
        
        # Simple chat display with auto-scrolling
        st.markdown(
            """
            <style>
            .scrollable-chat {
                max-height: 400px; 
                overflow-y: auto; 
                border: 1px solid #CCC;
                padding: 8px;
                margin-top: 10px;
                border-radius: 4px;
                background-color: #f9f9f9;
            }
            .message {
                padding: 0.5rem 1rem;
                border-radius: 15px;
                margin-bottom: 0.5rem;
                max-width: 80%;
                word-wrap: break-word;
            }
            .user-message {
                background-color: #e3f2fd;
                margin-left: auto;
                border-bottom-right-radius: 0;
            }
            .assistant-message {
                background-color: #f0f0f0;
                margin-right: auto;
                border-bottom-left-radius: 0;
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        st.markdown('<div class="scrollable-chat" id="chat-container">', unsafe_allow_html=True)
        for role, msg_text in st.session_state["chat_history"]:
            if role == "user":
                st.markdown(
                    f'<div class="message user-message"><strong>You:</strong> {msg_text}</div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f'<div class="message assistant-message"><strong>Bot:</strong> {msg_text}</div>',
                    unsafe_allow_html=True
                )
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Autoscroll
        st.markdown(
            """
            <script>
                function scrollToBottom() {
                    var container = document.getElementById('chat-container');
                    if(container) {
                        container.scrollTop = container.scrollHeight;
                    }
                }
                document.addEventListener('DOMContentLoaded', function() {
                    scrollToBottom();
                });
                const observer = new MutationObserver(function(mutations) {
                    scrollToBottom();
                });
                setTimeout(function() {
                    var container = document.getElementById('chat-container');
                    if(container) {
                        observer.observe(container, { childList: true });
                        scrollToBottom();
                    }
                }, 1000);
            </script>
            """,
            unsafe_allow_html=True
        )

if __name__ == '__main__':
    main()
